import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

# AISCR Brand Colors
NAVY = "1F4E78"
TURQUOISE = "2EC4B6"
GREEN = "00FF00"
YELLOW = "FFFF00"
RED = "FF0000"
LIGHT_GREY = "D3D3D3"

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Helper function to style header
def style_header(ws, row=1):
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.freeze_panes = ws[f'A{row+1}']
    ws.auto_filter.ref = ws.dimensions

# Helper function to add borders
def add_borders(ws, max_row, max_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).border = thin_border

# Helper function to auto-size columns
def auto_size_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================
# SHEET 1: Charter
# ============================================
ws1 = wb.create_sheet("Charter")
headers1 = ["Field", "Description"]
ws1.append(headers1)
style_header(ws1)
ws1.append(["Project Name", "AISCR Global PMO"])
ws1.append(["Project Manager", ""])
ws1.append(["Start Date", ""])
ws1.append(["End Date", ""])
ws1.append(["Objectives", ""])
ws1.append(["Success Criteria", ""])
add_borders(ws1, 10, 2)
auto_size_columns(ws1)

# ============================================
# SHEET 2: Scope
# ============================================
ws2 = wb.create_sheet("Scope")
headers2 = ["In Scope", "Out of Scope", "Constraints", "Assumptions"]
ws2.append(headers2)
style_header(ws2)
ws2.append(["Project management", "Non-PMO activities", "Budget limitations", "Stakeholder availability"])
ws2.append(["Volunteer coordination", "External projects", "Time constraints", "Resource availability"])
ws2.append(["Donor management", "Operational tasks", "Technology limitations", "Market stability"])
add_borders(ws2, 5, 4)
auto_size_columns(ws2)

# ============================================
# SHEET 3: WBS
# ============================================
ws3 = wb.create_sheet("WBS")
headers3 = ["WBS Level", "Task"]
ws3.append(headers3)
style_header(ws3)
ws3.append(["1.0", "Project Initiation"])
ws3.append(["1.1", "Charter Development"])
ws3.append(["1.2", "Stakeholder Identification"])
ws3.append(["2.0", "Planning"])
ws3.append(["2.1", "Scope Definition"])
ws3.append(["2.2", "Schedule Development"])
ws3.append(["3.0", "Execution"])
ws3.append(["3.1", "Resource Management"])
ws3.append(["3.2", "Risk Management"])
add_borders(ws3, 10, 2)
auto_size_columns(ws3)

# ============================================
# SHEET 4: Risk Register (Automated)
# ============================================
ws4 = wb.create_sheet("Risk Register")
headers4 = ["Risk ID", "Description", "Probability", "Impact", "Score", "Owner", "Mitigation", "Status"]
ws4.append(headers4)
style_header(ws4)

# Add sample data
ws4.append(["R001", "Resource unavailability", 3, 4, "=D2*C2", "PM", "Backup resources", "Yellow"])
ws4.append(["R002", "Budget overrun", 2, 5, "=D3*C3", "Finance", "Regular monitoring", "Red"])
ws4.append(["R003", "Scope creep", 4, 3, "=D4*C4", "PM", "Change control", "Yellow"])
ws4.append(["R004", "Technology failure", 2, 4, "=D5*C5", "IT", "Backup systems", "Green"])

# Status dropdown
status_dv = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=True)
ws4.add_data_validation(status_dv)
status_dv.add(f"H2:H100")

# Conditional formatting for Score
for row in range(2, 6):
    score_cell = ws4[f"E{row}"]
    score_cell.number_format = "0"
    
    # Conditional formatting rule
    ws4.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"E{row}<=5"], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws4.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"AND(E{row}>5,E{row}<=10)"], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws4.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"E{row}>10"], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws4, 6, 8)
auto_size_columns(ws4)

# ============================================
# SHEET 5: Stakeholders
# ============================================
ws5 = wb.create_sheet("Stakeholders")
headers5 = ["Stakeholder", "Role", "Influence", "Interest", "Engagement Strategy"]
ws5.append(headers5)
style_header(ws5)
ws5.append(["Executive Board", "Sponsor", "High", "High", "Regular updates"])
ws5.append(["Project Manager", "Leader", "High", "High", "Daily communication"])
ws5.append(["Volunteers", "Team", "Medium", "High", "Weekly meetings"])
ws5.append(["Donors", "External", "High", "Medium", "Quarterly reports"])
add_borders(ws5, 6, 5)
auto_size_columns(ws5)

# ============================================
# SHEET 6: Timeline + Live Gantt Chart
# ============================================
ws6 = wb.create_sheet("Timeline")
headers6 = ["Task", "Start Date", "End Date", "Duration", "Status"]
ws6.append(headers6)
style_header(ws6)

# Sample tasks with dates
today = datetime.now()
ws6.append(["Initiation", today, today + timedelta(days=30), "=C2-B2", "Not Started"])
ws6.append(["Planning", today + timedelta(days=31), today + timedelta(days=60), "=C3-B3", "In Progress"])
ws6.append(["Execution", today + timedelta(days=61), today + timedelta(days=120), "=C4-B4", "Not Started"])
ws6.append(["Closure", today + timedelta(days=121), today + timedelta(days=150), "=C5-B5", "Not Started"])

# Format date columns
for row in range(2, 6):
    ws6[f"B{row}"].number_format = "mm/dd/yyyy"
    ws6[f"C{row}"].number_format = "mm/dd/yyyy"
    ws6[f"D{row}"].number_format = "0"

# Status dropdown
status_dv2 = DataValidation(type="list", formula1='"Not Started,In Progress,Complete"', allow_blank=True)
ws6.add_data_validation(status_dv2)
status_dv2.add(f"E2:E100")

# Conditional formatting for Status
for row in range(2, 6):
    status_cell = ws6[f"E{row}"]
    ws6.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="Not Started"'], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws6.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="In Progress"'], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws6.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="Complete"'], fill=PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")))

add_borders(ws6, 6, 5)
auto_size_columns(ws6)

# ============================================
# SHEET 7: Budget (Automated)
# ============================================
ws7 = wb.create_sheet("Budget")
headers7 = ["Category", "Description", "Estimated Cost", "Actual Cost", "Variance", "Notes"]
ws7.append(headers7)
style_header(ws7)

ws7.append(["Personnel", "Staff salaries", 50000, 48000, "=D2-C2", ""])
ws7.append(["Technology", "Software licenses", 10000, 12000, "=D3-C3", ""])
ws7.append(["Operations", "Office supplies", 5000, 4500, "=D4-C4", ""])
ws7.append(["Marketing", "Promotional materials", 8000, 7500, "=D5-C5", ""])
ws7.append(["TOTAL", "", "=SUM(C2:C5)", "=SUM(D2:D5)", "=SUM(E2:E5)", ""])

# Format currency
for row in range(2, 7):
    for col in ['C', 'D', 'E']:
        ws7[f"{col}{row}"].number_format = "$#,##0.00"

# Conditional formatting for negative variance
for row in range(2, 6):
    ws7.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"E{row}<0"], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws7, 7, 6)
auto_size_columns(ws7)

# ============================================
# SHEET 8: Issues (Automated)
# ============================================
ws8 = wb.create_sheet("Issues")
headers8 = ["Issue ID", "Description", "Impact", "Owner", "Due Date", "Status", "Notes"]
ws8.append(headers8)
style_header(ws8)

today_issue = datetime.now()
ws8.append(["I001", "Resource shortage", "High", "PM", today_issue - timedelta(days=5), "Open", ""])
ws8.append(["I002", "Budget approval delay", "Medium", "Finance", today_issue + timedelta(days=10), "In Progress", ""])
ws8.append(["I003", "System downtime", "High", "IT", today_issue - timedelta(days=2), "Open", ""])
ws8.append(["I004", "Training needed", "Low", "HR", today_issue + timedelta(days=20), "Closed", ""])

# Format date column
for row in range(2, 6):
    ws8[f"E{row}"].number_format = "mm/dd/yyyy"

# Status dropdown
status_dv3 = DataValidation(type="list", formula1='"Open,In Progress,Closed"', allow_blank=True)
ws8.add_data_validation(status_dv3)
status_dv3.add(f"F2:F100")

# Highlight overdue issues
for row in range(2, 6):
    ws8.conditional_formatting.add(f"F{row}", 
        FormulaRule(formula=[f'AND(E{row}<TODAY(),F{row}<>"Closed")'], 
                   fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws8, 6, 7)
auto_size_columns(ws8)

# ============================================
# SHEET 9: Change Requests
# ============================================
ws9 = wb.create_sheet("Change Requests")
headers9 = ["Field", "Content"]
ws9.append(headers9)
style_header(ws9)
ws9.append(["Request ID", ""])
ws9.append(["Requestor", ""])
ws9.append(["Date", ""])
ws9.append(["Description", ""])
ws9.append(["Impact", ""])
ws9.append(["Status", ""])
add_borders(ws9, 8, 2)
auto_size_columns(ws9)

# ============================================
# SHEET 10: Lessons Learned
# ============================================
ws10 = wb.create_sheet("Lessons Learned")
headers10 = ["Aspect", "Details"]
ws10.append(headers10)
style_header(ws10)
ws10.append(["What Went Well", ""])
ws10.append(["What Could Improve", ""])
ws10.append(["Recommendations", ""])
ws10.append(["Best Practices", ""])
add_borders(ws10, 6, 2)
auto_size_columns(ws10)

# ============================================
# SHEET 11: Volunteers
# ============================================
ws11 = wb.create_sheet("Volunteers")
headers11 = ["Volunteer", "Skill", "Level", "Availability", "Notes"]
ws11.append(headers11)
style_header(ws11)
ws11.append(["John Doe", "Project Management", "Expert", "Full-time", ""])
ws11.append(["Jane Smith", "Data Analysis", "Intermediate", "Part-time", ""])
ws11.append(["Bob Johnson", "IT Support", "Expert", "Full-time", ""])
ws11.append(["Alice Brown", "Marketing", "Intermediate", "Part-time", ""])
add_borders(ws11, 6, 5)
auto_size_columns(ws11)

# ============================================
# SHEET 12: Assignments
# ============================================
ws12 = wb.create_sheet("Assignments")
headers12 = ["Volunteer", "Project", "Role", "Start Date", "Hours/Week", "Notes"]
ws12.append(headers12)
style_header(ws12)

ws12.append(["John Doe", "PMO System", "PM", today, 40, ""])
ws12.append(["Jane Smith", "Data Analysis", "Analyst", today, 20, ""])
ws12.append(["Bob Johnson", "IT Infrastructure", "IT Lead", today, 40, ""])
ws12.append(["Alice Brown", "Marketing Campaign", "Marketing Lead", today, 15, ""])

# Format date column
for row in range(2, 6):
    ws12[f"D{row}"].number_format = "mm/dd/yyyy"

add_borders(ws12, 6, 6)
auto_size_columns(ws12)

# ============================================
# SHEET 13: Contributions
# ============================================
ws13 = wb.create_sheet("Contributions")
headers13 = ["Volunteer", "Project", "Task", "Hours", "Date", "Review"]
ws13.append(headers13)
style_header(ws13)

ws13.append(["John Doe", "PMO System", "Planning", 40, today, 5])
ws13.append(["Jane Smith", "Data Analysis", "Analysis", 20, today, 4])
ws13.append(["Bob Johnson", "IT Infrastructure", "Setup", 30, today, 5])
ws13.append(["Alice Brown", "Marketing Campaign", "Design", 15, today, 4])
ws13.append(["John Doe", "PMO System", "Execution", 40, today + timedelta(days=7), 5])
ws13.append(["TOTAL", "", "", "=SUM(D2:D6)", "", ""])

# Format date column
for row in range(2, 7):
    ws13[f"E{row}"].number_format = "mm/dd/yyyy"

# Conditional formatting for low reviews
for row in range(2, 7):
    if row < 7:  # Exclude total row
        ws13.conditional_formatting.add(f"F{row}", 
            FormulaRule(formula=[f"F{row}<3"], 
                       fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws13, 7, 6)
auto_size_columns(ws13)

# ============================================
# SHEET 14: Portfolio (Automated)
# ============================================
ws14 = wb.create_sheet("Portfolio")
headers14 = ["Project", "Category", "Status", "Start Date", "End Date", "PM", "Analyst", "Priority"]
ws14.append(headers14)
style_header(ws14)

ws14.append(["PMO System", "Technology", "Green", today, today + timedelta(days=180), "John Doe", "Jane Smith", 1])
ws14.append(["Data Analysis", "Analytics", "Yellow", today, today + timedelta(days=120), "Jane Smith", "Bob Johnson", 2])
ws14.append(["IT Infrastructure", "Technology", "Green", today, today + timedelta(days=90), "Bob Johnson", "Alice Brown", 1])
ws14.append(["Marketing Campaign", "Marketing", "Red", today, today + timedelta(days=60), "Alice Brown", "John Doe", 3])

# Format date columns
for row in range(2, 6):
    ws14[f"D{row}"].number_format = "mm/dd/yyyy"
    ws14[f"E{row}"].number_format = "mm/dd/yyyy"

# Status dropdown
status_dv4 = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=True)
ws14.add_data_validation(status_dv4)
status_dv4.add(f"C2:C100")

# Priority dropdown
priority_dv = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
ws14.add_data_validation(priority_dv)
priority_dv.add(f"H2:H100")

# Conditional formatting for Status
for row in range(2, 6):
    ws14.conditional_formatting.add(f"C{row}", 
        FormulaRule(formula=[f'C{row}="Green"'], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws14.conditional_formatting.add(f"C{row}", 
        FormulaRule(formula=[f'C{row}="Yellow"'], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws14.conditional_formatting.add(f"C{row}", 
        FormulaRule(formula=[f'C{row}="Red"'], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws14, 6, 8)
auto_size_columns(ws14)

# ============================================
# SHEET 15: Risk Heat Map
# ============================================
ws15 = wb.create_sheet("Risk Heat Map")
headers15 = ["Likelihood", "Impact", "Score"]
ws15.append(headers15)
style_header(ws15)

# Create 5x5 heat grid
row_num = 2
for likelihood in range(1, 6):
    for impact in range(1, 6):
        score = likelihood * impact
        ws15.append([likelihood, impact, score])
        
        # Color scale based on score
        if score <= 5:
            fill_color = GREEN
        elif score <= 10:
            fill_color = YELLOW
        elif score <= 15:
            fill_color = "FFA500"  # Orange
        else:
            fill_color = RED
        
        ws15[f"C{row_num}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        row_num += 1

add_borders(ws15, 27, 3)
auto_size_columns(ws15)

# ============================================
# SHEET 16: Gantt Chart
# ============================================
ws16 = wb.create_sheet("Gantt Chart")
headers16 = ["Task", "Start Date", "End Date", "Duration", "Status"]
ws16.append(headers16)
style_header(ws16)

ws16.append(["Task 1", today, today + timedelta(days=30), "=C2-B2", "In Progress"])
ws16.append(["Task 2", today + timedelta(days=31), today + timedelta(days=60), "=C3-B3", "Not Started"])
ws16.append(["Task 3", today + timedelta(days=61), today + timedelta(days=90), "=C4-B4", "Not Started"])
ws16.append(["Task 4", today + timedelta(days=91), today + timedelta(days=120), "=C5-B5", "Not Started"])

# Format dates
for row in range(2, 6):
    ws16[f"B{row}"].number_format = "mm/dd/yyyy"
    ws16[f"C{row}"].number_format = "mm/dd/yyyy"
    ws16[f"D{row}"].number_format = "0"

# Status dropdown
status_dv5 = DataValidation(type="list", formula1='"Not Started,In Progress,Complete"', allow_blank=True)
ws16.add_data_validation(status_dv5)
status_dv5.add(f"E2:E100")

# Conditional formatting
for row in range(2, 6):
    ws16.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="Not Started"'], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws16.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="In Progress"'], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws16.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f'E{row}="Complete"'], fill=PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")))

add_borders(ws16, 6, 5)
auto_size_columns(ws16)

# ============================================
# SHEET 17: Stage-Gate Calculator
# ============================================
ws17 = wb.create_sheet("Stage-Gate Calculator")
headers17 = ["Gate", "Score", "Decision"]
ws17.append(headers17)
style_header(ws17)

ws17.append(["Initiation", 4, "=IF(B2>=3,\"PASS\",\"FAIL\")"])
ws17.append(["Planning", 3, "=IF(B3>=3,\"PASS\",\"FAIL\")"])
ws17.append(["Execution", 5, "=IF(B4>=3,\"PASS\",\"FAIL\")"])
ws17.append(["Closure", 2, "=IF(B5>=3,\"PASS\",\"FAIL\")"])

# Conditional formatting for Decision
for row in range(2, 6):
    ws17.conditional_formatting.add(f"C{row}", 
        FormulaRule(formula=[f'C{row}="PASS"'], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws17.conditional_formatting.add(f"C{row}", 
        FormulaRule(formula=[f'C{row}="FAIL"'], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws17, 6, 3)
auto_size_columns(ws17)

# ============================================
# SHEET 18: Donor Dashboard
# ============================================
ws18 = wb.create_sheet("Donor Dashboard")
headers18 = ["Donor", "Project", "Funding", "Status"]
ws18.append(headers18)
style_header(ws18)

ws18.append(["Donor A", "PMO System", 50000, "Green"])
ws18.append(["Donor B", "Data Analysis", 30000, "Yellow"])
ws18.append(["Donor C", "IT Infrastructure", 40000, "Green"])
ws18.append(["Donor D", "Marketing Campaign", 20000, "Red"])
ws18.append(["TOTAL", "", "=SUM(C2:C5)", ""])

# Format currency
for row in range(2, 6):
    ws18[f"C{row}"].number_format = "$#,##0.00"

# Status dropdown
status_dv6 = DataValidation(type="list", formula1='"Green,Yellow,Red"', allow_blank=True)
ws18.add_data_validation(status_dv6)
status_dv6.add(f"D2:D100")

# Conditional formatting for Status
for row in range(2, 6):
    ws18.conditional_formatting.add(f"D{row}", 
        FormulaRule(formula=[f'D{row}="Green"'], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws18.conditional_formatting.add(f"D{row}", 
        FormulaRule(formula=[f'D{row}="Yellow"'], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws18.conditional_formatting.add(f"D{row}", 
        FormulaRule(formula=[f'D{row}="Red"'], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

# Create bar chart
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Funding by Donor"
chart1.y_axis.title = "Funding ($)"
chart1.x_axis.title = "Donor"

data = openpyxl.chart.Reference(ws18, min_col=3, min_row=1, max_row=5)
cats = openpyxl.chart.Reference(ws18, min_col=1, min_row=2, max_row=5)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 10
chart1.width = 15

ws18.add_chart(chart1, "F2")

add_borders(ws18, 6, 4)
auto_size_columns(ws18)

# ============================================
# SHEET 19: Membership Dashboard
# ============================================
ws19 = wb.create_sheet("Membership Dashboard")
headers19 = ["Member Type", "Count", "Growth %", "Notes"]
ws19.append(headers19)
style_header(ws19)

ws19.append(["Regular", 150, "=((B2-120)/120)*100", ""])
ws19.append(["Premium", 50, "=((B3-40)/40)*100", ""])
ws19.append(["VIP", 20, "=((B4-15)/15)*100", ""])
ws19.append(["TOTAL", "=SUM(B2:B4)", "", ""])

# Format percentage
for row in range(2, 5):
    ws19[f"C{row}"].number_format = "0.00%"

# Create line chart
chart2 = LineChart()
chart2.title = "Membership Growth Trends"
chart2.style = 13
chart2.y_axis.title = "Count"
chart2.x_axis.title = "Member Type"

data2 = openpyxl.chart.Reference(ws19, min_col=2, min_row=1, max_row=4)
cats2 = openpyxl.chart.Reference(ws19, min_col=1, min_row=2, max_row=4)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height = 10
chart2.width = 15

ws19.add_chart(chart2, "F2")

add_borders(ws19, 5, 4)
auto_size_columns(ws19)

# ============================================
# SHEET 20: Volunteer Scorecard
# ============================================
ws20 = wb.create_sheet("Volunteer Scorecard")
headers20 = ["Volunteer", "Tasks Completed", "Hours", "Quality Rating", "Performance Score"]
ws20.append(headers20)
style_header(ws20)

ws20.append(["John Doe", 10, 200, 5, "=(B2*0.4)+(C2*0.3)+(D2*0.3)"])
ws20.append(["Jane Smith", 8, 150, 4, "=(B3*0.4)+(C3*0.3)+(D3*0.3)"])
ws20.append(["Bob Johnson", 12, 180, 5, "=(B4*0.4)+(C4*0.3)+(D4*0.3)"])
ws20.append(["Alice Brown", 6, 100, 3, "=(B5*0.4)+(C5*0.3)+(D5*0.3)"])

# Format Performance Score
for row in range(2, 6):
    ws20[f"E{row}"].number_format = "0.00"

# Conditional formatting for Performance Score
for row in range(2, 6):
    ws20.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"E{row}>=80"], fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")))
    ws20.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"AND(E{row}>=50,E{row}<80)"], fill=PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")))
    ws20.conditional_formatting.add(f"E{row}", 
        FormulaRule(formula=[f"E{row}<50"], fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid")))

add_borders(ws20, 6, 5)
auto_size_columns(ws20)

# ============================================
# SHEET 21: PMO KPIs
# ============================================
ws21 = wb.create_sheet("PMO KPIs")
headers21 = ["Metric", "Value"]
ws21.append(headers21)
style_header(ws21)

ws21.append(["Total Projects", "=COUNTA(Portfolio!A2:A100)"])
ws21.append(["Active Projects", "=COUNTIF(Portfolio!C2:C100,\"Green\")"])
ws21.append(["At Risk Projects", "=COUNTIF(Portfolio!C2:C100,\"Yellow\")"])
ws21.append(["Critical Projects", "=COUNTIF(Portfolio!C2:C100,\"Red\")"])
ws21.append(["Total Funding", "=SUM('Donor Dashboard'!C2:C5)"])
ws21.append(["Membership Growth", "=AVERAGE('Membership Dashboard'!C2:C4)"])
ws21.append(["Average Volunteer Score", "=AVERAGE('Volunteer Scorecard'!E2:E5)"])

# Format values
ws21["B3"].number_format = "0"
ws21["B4"].number_format = "0"
ws21["B5"].number_format = "0"
ws21["B6"].number_format = "$#,##0.00"
ws21["B7"].number_format = "0.00%"
ws21["B8"].number_format = "0.00"

add_borders(ws21, 9, 2)
auto_size_columns(ws21)

# ============================================
# SHEET 22: Dashboard (Fully Automated)
# ============================================
ws22 = wb.create_sheet("Dashboard")
ws22.sheet_view.showGridLines = False

# Title
title_cell = ws22["A1"]
title_cell.value = "AISCR GLOBAL PMO DASHBOARD"
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws22.merge_cells("A1:H1")
ws22.row_dimensions[1].height = 30

# KPI Cards Section
row = 3
kpi_labels = ["Total Projects", "Active (Green)", "At Risk (Yellow)", "Critical (Red)", 
              "Total Funding", "Membership Growth", "Average Volunteer Score"]

for i, label in enumerate(kpi_labels):
    col = (i % 4) * 2 + 1
    if i > 0 and i % 4 == 0:
        row += 3
    
    # Label
    label_cell = ws22.cell(row, col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal='center')
    
    # Value
    value_cell = ws22.cell(row + 1, col)
    if label == "Total Projects":
        value_cell.value = "='PMO KPIs'!B2"
    elif label == "Active (Green)":
        value_cell.value = "='PMO KPIs'!B3"
    elif label == "At Risk (Yellow)":
        value_cell.value = "='PMO KPIs'!B4"
    elif label == "Critical (Red)":
        value_cell.value = "='PMO KPIs'!B5"
    elif label == "Total Funding":
        value_cell.value = "='PMO KPIs'!B6"
        value_cell.number_format = "$#,##0.00"
    elif label == "Membership Growth":
        value_cell.value = "='PMO KPIs'!B7"
        value_cell.number_format = "0.00%"
    elif label == "Average Volunteer Score":
        value_cell.value = "='PMO KPIs'!B8"
        value_cell.number_format = "0.00"
    
    value_cell.font = Font(bold=True, size=14, color="FFFFFF")
    value_cell.fill = PatternFill(start_color=TURQUOISE, end_color=TURQUOISE, fill_type="solid")
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    value_cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws22.column_dimensions[get_column_letter(col)].width = 18
    ws22.row_dimensions[row + 1].height = 40

# Charts Section - Start at row 12
chart_row = 12

# Portfolio Status Chart
chart3 = BarChart()
chart3.type = "col"
chart3.style = 10
chart3.title = "Portfolio Status"
chart3.y_axis.title = "Count"

data3 = openpyxl.chart.Reference(ws14, min_col=3, min_row=1, max_row=5)
cats3 = openpyxl.chart.Reference(ws14, min_col=1, min_row=2, max_row=5)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.height = 8
chart3.width = 12

ws22.add_chart(chart3, f"A{chart_row}")

# Donor Funding Chart (copy from Donor Dashboard)
chart4 = BarChart()
chart4.type = "col"
chart4.style = 10
chart4.title = "Donor Funding"
chart4.y_axis.title = "Funding ($)"

data4 = openpyxl.chart.Reference(ws18, min_col=3, min_row=1, max_row=5)
cats4 = openpyxl.chart.Reference(ws18, min_col=1, min_row=2, max_row=5)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.height = 8
chart4.width = 12

ws22.add_chart(chart4, f"F{chart_row}")

# Membership Growth Chart (copy from Membership Dashboard)
chart5 = LineChart()
chart5.title = "Membership Growth"
chart5.style = 13
chart5.y_axis.title = "Growth %"

data5 = openpyxl.chart.Reference(ws19, min_col=3, min_row=1, max_row=4)
cats5 = openpyxl.chart.Reference(ws19, min_col=1, min_row=2, max_row=4)
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
chart5.height = 8
chart5.width = 12

ws22.add_chart(chart5, f"A{chart_row + 10}")

# Risk Heat Map Snapshot
ws22.cell(chart_row + 10, 6).value = "Risk Heat Map Summary"
ws22.cell(chart_row + 10, 6).font = Font(bold=True, size=12)
ws22.cell(chart_row + 11, 6).value = "Low Risk (≤5):"
ws22.cell(chart_row + 11, 7).value = "=COUNTIF('Risk Heat Map'!C2:C26,\"<=5\")"
ws22.cell(chart_row + 12, 6).value = "Medium Risk (6-10):"
ws22.cell(chart_row + 12, 7).value = "=COUNTIFS('Risk Heat Map'!C2:C26,\">5\",'Risk Heat Map'!C2:C26,\"<=10\")"
ws22.cell(chart_row + 13, 6).value = "High Risk (>10):"
ws22.cell(chart_row + 13, 7).value = "=COUNTIF('Risk Heat Map'!C2:C26,\">10\")"

auto_size_columns(ws22)

# Save workbook
wb.save("AISCR_PMO_Full_Automated_System.xlsx")
print("Excel file created successfully: AISCR_PMO_Full_Automated_System.xlsx")

